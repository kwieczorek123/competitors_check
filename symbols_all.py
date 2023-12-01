from MetaTrader5 import initialize, shutdown, symbol_info_tick, symbol_info, symbols_get, terminal_info
import openpyxl
import numpy as np
from openpyxl.styles import PatternFill
import re

# Initialize the MetaTrader 5 terminal
if not initialize():
    print("Initialization has failed")
    shutdown()


def get_usd_rate(quote_currency):
    if quote_currency == 'USD':
        return 1
    else:
        usd_tick = symbol_info_tick(f"{quote_currency}USD")
        if usd_tick:
            return usd_tick.ask
    return None


# Get all available symbols and their categories
all_symbols = symbols_get()
symbols = [symbol.name for symbol in all_symbols]  # List of all symbol names
symbol_categories = {symbol.name: symbol.path for symbol in all_symbols}  # Dictionary of symbol names to their categories


tick_data = {}

for symbol in symbols:
    tick = symbol_info_tick(symbol)
    info = symbol_info(symbol)

    if not tick or not info:
        print(f"Could not retrieve tick or info for {symbol}.")
        continue

    if tick and info:
        # Determine the currency pair or other instrument
        if symbol[-3:] in ['USD', 'EUR', 'GBP', 'JPY', 'CAD', 'AUD', 'NZD', 'CHF']:
            quote_currency = symbol[-3:]
        else:
            quote_currency = info.currency_profit  # or info.currency_margin

        usd_rate = get_usd_rate(quote_currency)

        if not usd_rate:
            continue

        spread = tick.ask - tick.bid  # This is the raw spread
        spread_in_points = spread * (10 ** info.digits)  # adjust raw spread to points

        spread_in_quote = spread  # Use raw spread for calculations
        spread_in_usd_per_contract = spread_in_quote * usd_rate * info.trade_contract_size
        spread_in_usd_per_contract = np.round(spread_in_usd_per_contract, 2)

        min_volume = info.volume_min
        max_volume = info.volume_max
        limit_volume = info.volume_limit
        contract_size = info.trade_contract_size
        margin = info.margin_initial

        swap_long = info.swap_long if info.swap_mode != 0 else 0
        swap_short = info.swap_short if info.swap_mode != 0 else 0
        limit_volume = info.volume_limit
        trade_calc_mode = info.trade_calc_mode
        underlying_lot_amount_usd = tick.bid * contract_size * usd_rate
        underlying_lot_amount_usd = np.round(underlying_lot_amount_usd, 2)

        tick_data[symbol] = {
            "price": tick.ask,
            "spread": spread,
            "spread_in_points": spread_in_points,  # added to the dictionary
            "contract_size": contract_size,
            "min_volume": min_volume,
            "max_volume": max_volume,
            "limit_volume": limit_volume,
            "margin": margin,
            "quote_currency": quote_currency,
            "usd_rate": usd_rate,
            "spread_in_usd_per_contract": spread_in_usd_per_contract,
            "swap_long": swap_long,
            "swap_short": swap_short,
            "type": "",  # Empty by default
            "commission": "",  # Empty by default
            "underlying_lot_amount_usd": underlying_lot_amount_usd,
            "ib_commission_per_lot": "",  # Empty by default
            "trade_calc_mode": trade_calc_mode
        }

        # Add category information to the dictionary
        tick_data[symbol]['category'] = symbol_categories.get(symbol, 'Unknown')

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Write headers to Excel
headers = ["Symbol", "Category", "Type", "Price", "Spread", "Spread in Points", "Spread in USD", "Contract Size", "Min Volume",
           "Max Volume", "Limit Volume", "Margin Buy", "Trade Calculation Mode", "Quote Currency", "USD rate",
           "Commission", "Swap Long", "Swap Short", "Underlying lot amount USD",
           "IB Commission per lot"]
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num).value = header

# Create a yellow fill for background color
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Write data to Excel
row_num = 2
for symbol, data in tick_data.items():
    # Check for symbols where we want an empty row before them
    if symbol in ["AUS200", "XBRUSD", "XAGUSD"]:
        row_num += 1  # Increment row_num to leave an empty row

    ws.cell(row=row_num, column=1).value = symbol
    ws.cell(row=row_num, column=2).value = data['category']
    ws.cell(row=row_num, column=3).value = data["type"]
    ws.cell(row=row_num, column=4).value = data["price"]
    ws.cell(row=row_num, column=5).value = data["spread"]
    ws.cell(row=row_num, column=6).value = data["spread_in_points"]
    ws.cell(row=row_num, column=7).value = data["spread_in_usd_per_contract"]
    ws.cell(row=row_num, column=8).value = data["contract_size"]
    ws.cell(row=row_num, column=9).value = data["min_volume"]
    ws.cell(row=row_num, column=10).value = data["max_volume"]
    ws.cell(row=row_num, column=11).value = data["limit_volume"]
    ws.cell(row=row_num, column=12).value = data["margin"]
    ws.cell(row=row_num, column=13).value = data["trade_calc_mode"]
    ws.cell(row=row_num, column=14).value = data["quote_currency"]
    ws.cell(row=row_num, column=15).value = data["usd_rate"]
    ws.cell(row=row_num, column=16).value = data["commission"]
    ws.cell(row=row_num, column=17).value = data["swap_long"]
    ws.cell(row=row_num, column=18).value = data["swap_short"]
    ws.cell(row=row_num, column=19).value = data["underlying_lot_amount_usd"]
    ws.cell(row=row_num, column=20).value = data["ib_commission_per_lot"]

    # Apply the yellow fill to the appropriate cells
    ws.cell(row=row_num, column=3).fill = yellow_fill  # Type column
    ws.cell(row=row_num, column=16).fill = yellow_fill  # Commission column
    ws.cell(row=row_num, column=20).fill = yellow_fill  # IB Commission per lot column

    row_num += 1

for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:  # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)  # adding a little extra space
    ws.column_dimensions[column[0].column_letter].width = adjusted_width

# Retrieve broker's name from MT5 terminal info
broker_info = terminal_info()
broker_name = broker_info.name
print(broker_name)

# Sanitize the broker's name to ensure it's a valid filename (remove any invalid characters)
valid_filename = re.sub(r'[\\/*?:"<>|]',"", broker_name)

# Create the filename with the broker's name
filename = f"{valid_filename}_all_symbols.xlsx"

# Save the workbook with the new filename
wb.save(filename)

# Shutdown the connection to MetaTrader 5
shutdown()
