from MetaTrader5 import initialize, shutdown, symbol_info_tick, symbol_info, terminal_info, symbols_get
import openpyxl
import numpy as np
import re
import logging
import mappings


# Initialize the MetaTrader 5 terminal
if not initialize():
    print("Initialization has failed")
    shutdown()


# Configure logging
logging.basicConfig(level=logging.DEBUG, filename='mt5_symbols_log.txt', filemode='w',
                    format='%(name)s - %(levelname)s - %(message)s')

# Retrieve broker's name from MT5 terminal info
broker_info = terminal_info()
broker_name = broker_info.name
print(broker_name)

# Select the appropriate mappings based on the broker's name
if broker_name.lower() == 'oanda metatrader 5':
    symbols = mappings.symbols_oanda
    symbols_digits = mappings.symbols_digits_oanda
elif broker_name.lower() == 'avatrade':
    symbols = mappings.symbols_avatrade
    symbols_digits = mappings.symbols_digits_avatrade
# Add more elif statements for other brokers
else:
    symbols = mappings.symbols  # Default mappings
    symbols_digits = mappings.symbols_digits


def get_profit_currency(symbol):
    # Retrieve symbol info
    info = symbol_info(symbol)
    if info:
        return info.currency_profit
    return None


def find_matching_symbol(currency, symbols_list):
    # Find a symbol in the list that pairs the currency with USD
    for sym in symbols_list:
        if currency + 'USD' in sym or 'USD' + currency in sym:
            return sym
    return None


def get_usd_rate(symbol, symbols_list):
    profit_currency = get_profit_currency(symbol)
    print(f"Profit Currency for {symbol}: {profit_currency}")  # Debug print

    if profit_currency == 'USD':
        return 1

    matched_symbol = find_matching_symbol(profit_currency, symbols_list)
    print(f"Matched Symbol for {symbol}: {matched_symbol}")  # Debug print

    if matched_symbol:
        tick = symbol_info_tick(matched_symbol)
        print(f"Tick data for {matched_symbol}: Bid = {tick.bid}, Ask = {tick.ask}")  # Debug print

        if tick:
            if matched_symbol.startswith('USD'):
                usd_rate = 1 / tick.bid
            else:
                usd_rate = tick.ask

            print(f"USD Rate for {symbol}: {usd_rate}")  # Debug print
            return usd_rate

    return None


def adjust_value_by_digits(value, actual_digits, reference_digits):
    if not actual_digits:
        actual_digits = 0
    else:
        actual_digits = int(actual_digits)

    if not reference_digits:
        reference_digits = 0
    else:
        reference_digits = int(reference_digits)

    difference = actual_digits - reference_digits
    factor = 10 ** abs(difference)

    if difference < 0:
        return value * factor
    elif difference > 0:
        return value / factor
    else:
        return value


# Retrieve all available symbols from MT5
available_symbols = symbols_get()

# Extract the symbol names into a list
broker_symbol_names = [symbol.name for symbol in available_symbols]

# Print out the broker symbol names
print(broker_symbol_names)


tick_data = {}

for symbol in symbols:
    tick = symbol_info_tick(symbol)
    info = symbol_info(symbol)

    if not tick or not info:
        if not tick:
            logging.error(f"Could not retrieve tick data for {symbol}.")
        if not info:
            logging.error(f"Could not retrieve symbol info for {symbol}.")
        # Attempt to get more detailed information about the symbol
        symbol_details = symbol_info(symbol)
        if symbol_details:
            logging.info(f"Symbol details for {symbol}: {symbol_details}")
        else:
            logging.error(f"Symbol details for {symbol} could not be retrieved.")
        continue

    if tick and info:
        # Determine the currency pair or other instrument
        if symbol[-3:] in ['USD', 'EUR', 'GBP', 'JPY', 'CAD', 'AUD', 'NZD', 'CHF']:
            quote_currency = symbol[-3:]
        else:
            quote_currency = info.currency_profit  # or info.currency_margin

        usd_rate = get_usd_rate(symbol, symbols)

        if not usd_rate:
            continue

        digits = info.digits
        spread = tick.ask - tick.bid  # This is the raw spread

        if digits != 0:
            spread_in_points = spread * (10 ** digits)  # adjust raw spread to points
        else:
            spread_in_points = spread  # or any other default value

        spread_in_quote = spread  # Use raw spread for calculations
        spread_in_usd_per_contract = spread_in_quote * usd_rate * info.trade_contract_size
        spread_in_usd_per_contract = np.round(spread_in_usd_per_contract, 2)

        min_volume = info.volume_min
        max_volume = info.volume_max
        limit_volume = info.volume_limit
        contract_size = info.trade_contract_size
        margin = info.margin_initial

        swap_mode = info.swap_mode
        swap_long = info.swap_long if info.swap_mode != 0 else 0
        swap_short = info.swap_short if info.swap_mode != 0 else 0
        limit_volume = info.volume_limit
        trade_calc_mode = info.trade_calc_mode
        underlying_lot_amount_usd = tick.bid * contract_size * usd_rate
        underlying_lot_amount_usd = np.round(underlying_lot_amount_usd, 2)

        tick_data[symbol] = {
            "price": tick.ask,
            "digits": digits,
            "spread": spread,
            "spread_in_points": spread_in_points,
            "contract_size": contract_size,
            "min_volume": min_volume,
            "max_volume": max_volume,
            "limit_volume": limit_volume,
            "margin": margin,
            "quote_currency": quote_currency,
            "usd_rate": usd_rate,
            "spread_in_usd_per_contract": spread_in_usd_per_contract,
            "swap_mode": swap_mode,
            "swap_long": swap_long,
            "swap_short": swap_short,
            "commission": "",  # Empty by default
            "underlying_lot_amount_usd": underlying_lot_amount_usd,
            "trade_calc_mode": trade_calc_mode
        }

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Write headers to Excel
headers = ["Symbol", "Type", "Price", "Digits", "Spread", "Spread in Points", "Spread in Octa Points", "Spread in USD",
           "Contract Size", "Min Volume", "Max Volume", "Limit Volume", "Margin Buy", "Trade Calculation Mode",
           "Quote Currency", "USD rate", "Commission", "Swap mode", "Swap Long", "Swap Short",
           "Swap Long in Octa Points", "Swap Short in Octa Points", "Underlying lot amount USD",
           "IB Commission per lot"]

for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num).value = header

# Write data to Excel
row_num = 2
for symbol in symbols:  # We loop through symbols list to make sure all symbols are captured

    data = tick_data.get(symbol, {})  # Use a default empty dictionary if symbol data is missing

    ws.cell(row=row_num, column=1).value = symbol
    ws.cell(row=row_num, column=2).value = data.get("type", "")
    ws.cell(row=row_num, column=3).value = data.get("price", "")
    ws.cell(row=row_num, column=4).value = data.get("digits", "")
    ws.cell(row=row_num, column=5).value = data.get("spread", "")
    ws.cell(row=row_num, column=6).value = data.get("spread_in_points", "")
    # Adjusting spread_in_points
    actual_digits = data.get("digits", "")
    reference_digits = symbols_digits.get(symbol, actual_digits)  # default to actual_digits if symbol not found
    ws.cell(row=row_num, column=7).value = adjust_value_by_digits(data.get("spread_in_points", 0), actual_digits,
                                                                  reference_digits)

    ws.cell(row=row_num, column=8).value = data.get("spread_in_usd_per_contract", "")
    ws.cell(row=row_num, column=9).value = data.get("contract_size", "")
    ws.cell(row=row_num, column=10).value = data.get("min_volume", "")
    ws.cell(row=row_num, column=11).value = data.get("max_volume", "")
    ws.cell(row=row_num, column=12).value = data.get("limit_volume", "")
    ws.cell(row=row_num, column=13).value = data.get("margin", "")
    ws.cell(row=row_num, column=14).value = data.get("trade_calc_mode", "")
    ws.cell(row=row_num, column=15).value = data.get("quote_currency", "")
    ws.cell(row=row_num, column=16).value = data.get("usd_rate", "")
    ws.cell(row=row_num, column=17).value = data.get("commission", "")
    ws.cell(row=row_num, column=18).value = data.get("swap_mode", "")
    ws.cell(row=row_num, column=19).value = data.get("swap_long", "")
    ws.cell(row=row_num, column=20).value = data.get("swap_short", "")
    # Adjusting swap values
    ws.cell(row=row_num, column=21).value = adjust_value_by_digits(data.get("swap_long", 0), actual_digits,
                                                                   reference_digits)
    ws.cell(row=row_num, column=22).value = adjust_value_by_digits(data.get("swap_short", 0), actual_digits,
                                                                   reference_digits)

    ws.cell(row=row_num, column=23).value = data.get("underlying_lot_amount_usd", "")
    ws.cell(row=row_num, column=24).value = data.get("ib_commission_per_lot", "")

    row_num += 1

for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:  # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)  # adding a little extra space
    ws.column_dimensions[column[0].column_letter].width = adjusted_width

# Sanitize the broker's name to ensure it's a valid filename (remove any invalid characters)
valid_filename = re.sub(r'[\\/*?:"<>|]',"", broker_name)

# Create the filename with the broker's name
filename = f"{valid_filename}_symbols_info.xlsx"

# Save the workbook with the new filename
wb.save(filename)

logging.info("Script completed")

# Shutdown the connection to MetaTrader 5
shutdown()
