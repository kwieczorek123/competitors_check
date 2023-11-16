from MetaTrader5 import initialize, shutdown, symbol_info_tick, symbol_info, terminal_info, symbols_get
import openpyxl
import numpy as np
import re
import logging


# Initialize the MetaTrader 5 terminal
if not initialize():
    print("Initialization has failed")
    shutdown()


# Configure logging
logging.basicConfig(level=logging.DEBUG, filename='mt5_symbols_log.txt', filemode='w',
                    format='%(name)s - %(levelname)s - %(message)s')


def get_usd_rate(quote_currency):
    if quote_currency == 'USD':
        return 1
    else:
        # Try the direct quote first
        direct_tick = symbol_info_tick(f"{quote_currency}USD")
        if direct_tick:
            return direct_tick.ask
        else:
            # If the direct quote fails, try the inverse quote
            inverse_tick = symbol_info_tick(f"USD{quote_currency}")
            if inverse_tick:
                return 1 / inverse_tick.bid  # use bid price to convert from USD to the quote currency
    return None


def adjust_value_by_digits(value, actual_digits, reference_digits):
    if not actual_digits:
        actual_digits = 0
    else:
        actual_digits = int(actual_digits)

    if not reference_digits:
        reference_digits = 0
    else:
        reference_digits = int(reference_digits)

    difference = actual_digits - reference_digits
    factor = 10 ** abs(difference)

    if difference < 0:
        return value * factor
    elif difference > 0:
        return value / factor
    else:
        return value


symbols = [
    "XAUUSD", "XAGUSD", "EURUSD", "GBPUSD", "USDJPY", "GBPJPY", "USDCHF", "EURGBP",
    "AUDUSD", "NZDUSD", "USDCAD", "AUDCAD", "AUDJPY", "CADJPY", "CHFJPY", "EURAUD",
    "EURJPY", "GBPAUD", "NAS100", "US30", "SPX500", "JPN225", "GER40", "XBRUSD",
    "XTIUSD", "XNGUSD", "BTCUSD", "ETHUSD", "LTCUSD", "XRPUSD", "BCHUSD",
    "USDPHP", "USDBRL", "USDTHB"
]

symbols_digits = {
    "EURGBP": 5,  "EURJPY": 3,  "EURCHF": 5,  "EURAUD": 5,  "EURNZD": 5,  "EURCAD": 5,
    "GBPJPY": 3,  "GBPCHF": 5,  "GBPAUD": 5,  "GBPNZD": 5,  "GBPCAD": 5,  "CHFJPY": 3,
    "AUDJPY": 3,  "AUDCHF": 5,  "AUDNZD": 5,  "AUDCAD": 5,  "NZDJPY": 3,  "NZDCHF": 5,
    "NZDCAD": 5,  "CADJPY": 3,  "CADCHF": 5,  "USDMXN": 4,  "EURMXN": 4,  "GBPMXN": 4,
    "EURZAR": 5,  "USDZAR": 5,  "GBPZAR": 5,  "ZARJPY": 3,  "USDHKD": 5,  "USDSEK": 5,
    "USDSGD": 5,  "EURUSD": 5,  "GBPUSD": 5,  "USDJPY": 3,  "USDCHF": 5,  "AUDUSD": 5,
    "NZDUSD": 5,  "USDCAD": 5,  "XAUUSD": 2,  "XAGUSD": 3,  "NAS100": 1,  "US30": 1,
    "SPX500": 1,  "JPN225": 0,  "GER40": 1,   "XBRUSD": 2,  "XTIUSD": 2,  "XNGUSD": 3,
    "BTCUSD": 2,  "ETHUSD": 2,  "LTCUSD": 2,  "XRPUSD": 5,  "BCHUSD": 2,
    "USDPHP": 2,  "USDBRL": 4,  "USDTHB": 3
}


# Retrieve all available symbols from MT5
available_symbols = symbols_get()

# Extract the symbol names into a list
broker_symbol_names = [symbol.name for symbol in available_symbols]

# Print out the broker symbol names
print(broker_symbol_names)


tick_data = {}

for symbol in symbols:
    tick = symbol_info_tick(symbol)
    info = symbol_info(symbol)

    if not tick or not info:
        if not tick:
            logging.error(f"Could not retrieve tick data for {symbol}.")
        if not info:
            logging.error(f"Could not retrieve symbol info for {symbol}.")
        # Attempt to get more detailed information about the symbol
        symbol_details = symbol_info(symbol)
        if symbol_details:
            logging.info(f"Symbol details for {symbol}: {symbol_details}")
        else:
            logging.error(f"Symbol details for {symbol} could not be retrieved.")
        continue

    if tick and info:
        # Determine the currency pair or other instrument
        if symbol[-3:] in ['USD', 'EUR', 'GBP', 'JPY', 'CAD', 'AUD', 'NZD', 'CHF']:
            quote_currency = symbol[-3:]
        else:
            quote_currency = info.currency_profit  # or info.currency_margin

        usd_rate = get_usd_rate(quote_currency)

        if not usd_rate:
            continue

        digits = info.digits
        spread = tick.ask - tick.bid  # This is the raw spread

        if digits != 0:
            spread_in_points = spread * (10 ** digits)  # adjust raw spread to points
        else:
            spread_in_points = spread  # or any other default value

        spread_in_quote = spread  # Use raw spread for calculations
        spread_in_usd_per_contract = spread_in_quote * usd_rate * info.trade_contract_size
        spread_in_usd_per_contract = np.round(spread_in_usd_per_contract, 2)

        contract_size = info.trade_contract_size
        margin = info.margin_initial

        swap_mode = info.swap_mode
        swap_long = info.swap_long if info.swap_mode != 0 else 0
        swap_short = info.swap_short if info.swap_mode != 0 else 0
        limit_volume = info.volume_limit
        trade_calc_mode = info.trade_calc_mode
        underlying_lot_amount_usd = tick.bid * contract_size * usd_rate
        underlying_lot_amount_usd = np.round(underlying_lot_amount_usd, 2)

        tick_data[symbol] = {
            "price": tick.ask,
            "digits": digits,
            "spread": spread,
            "spread_in_points": spread_in_points,
            "contract_size": contract_size,
            "margin": margin,
            "quote_currency": quote_currency,
            "usd_rate": usd_rate,
            "spread_in_usd_per_contract": spread_in_usd_per_contract,
            "swap_mode": swap_mode,
            "swap_long": swap_long,
            "swap_short": swap_short,
            "commission": "",  # Empty by default
            "underlying_lot_amount_usd": underlying_lot_amount_usd,
            "trade_calc_mode": trade_calc_mode
        }

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Write headers to Excel
headers = ["Symbol", "Price", "Digits", "Spread", "Spread in Points", "Spread in Octa Points", "Spread in USD",
           "Contract Size", "Margin Buy", "Trade Calculation Mode", "Quote Currency", "USD rate", "Commission",
           "Swap mode", "Swap Long", "Swap Short", "Swap Long in Octa Points", "Swap Short in Octa Points",
           "Underlying lot amount USD"]

for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num).value = header

# Write data to Excel
row_num = 2
for symbol in symbols:  # We loop through symbols list to make sure all symbols are captured

    # Check for symbols where we want an empty row before them
    if symbol in ["EURUSD", "NAS100", "XBRUSD", "BTCUSD", "USDPHP"]:
        row_num += 1  # Increment row_num to leave an empty row

    data = tick_data.get(symbol, {})  # Use a default empty dictionary if symbol data is missing

    ws.cell(row=row_num, column=1).value = symbol
    ws.cell(row=row_num, column=2).value = data.get("price", "")
    ws.cell(row=row_num, column=3).value = data.get("digits", "")
    ws.cell(row=row_num, column=4).value = data.get("spread", "")
    ws.cell(row=row_num, column=5).value = data.get("spread_in_points", "")
    # Adjusting spread_in_points
    actual_digits = data.get("digits", "")
    reference_digits = symbols_digits.get(symbol, actual_digits)  # default to actual_digits if symbol not found
    ws.cell(row=row_num, column=6).value = adjust_value_by_digits(data.get("spread_in_points", 0), actual_digits,
                                                                  reference_digits)
    ws.cell(row=row_num, column=7).value = data.get("spread_in_usd_per_contract", "")
    ws.cell(row=row_num, column=8).value = data.get("contract_size", "")
    ws.cell(row=row_num, column=9).value = data.get("margin", "")
    ws.cell(row=row_num, column=10).value = data.get("trade_calc_mode", "")
    ws.cell(row=row_num, column=11).value = data.get("quote_currency", "")
    ws.cell(row=row_num, column=12).value = data.get("usd_rate", "")
    ws.cell(row=row_num, column=13).value = data.get("commission", "")
    ws.cell(row=row_num, column=14).value = data.get("swap_mode", "")
    ws.cell(row=row_num, column=15).value = data.get("swap_long", "")
    ws.cell(row=row_num, column=16).value = data.get("swap_short", "")
    # Adjusting swap values
    ws.cell(row=row_num, column=17).value = adjust_value_by_digits(data.get("swap_long", 0), actual_digits,
                                                                   reference_digits)
    ws.cell(row=row_num, column=18).value = adjust_value_by_digits(data.get("swap_short", 0), actual_digits,
                                                                   reference_digits)

    ws.cell(row=row_num, column=19).value = data.get("underlying_lot_amount_usd", "")

    row_num += 1

for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:  # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)  # adding a little extra space
    ws.column_dimensions[column[0].column_letter].width = adjusted_width

# Retrieve broker's name from MT5 terminal info
broker_info = terminal_info()
broker_name = broker_info.name

# Sanitize the broker's name to ensure it's a valid filename (remove any invalid characters)
valid_filename = re.sub(r'[\\/*?:"<>|]',"", broker_name)

# Create the filename with the broker's name
filename = f"{valid_filename}_symbols_info.xlsx"

# Save the workbook with the new filename
wb.save(filename)

logging.info("Script completed")

# Shutdown the connection to MetaTrader 5
shutdown()
