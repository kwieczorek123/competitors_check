from MetaTrader5 import initialize, shutdown, symbol_info_tick, symbol_info
import openpyxl
import numpy as np
from openpyxl.styles import PatternFill

# Initialize the MetaTrader 5 terminal
if not initialize():
    print("Initialization has failed")
    shutdown()


def get_usd_rate(quote_currency):
    if quote_currency == 'USD':
        return 1
    else:
        usd_tick = symbol_info_tick(f"{quote_currency}USD")
        if usd_tick:
            return usd_tick.ask
    return None


def adjust_value_by_digits(value, actual_digits, reference_digits):
    if not actual_digits:
        actual_digits = 0
    else:
        actual_digits = int(actual_digits)

    if not reference_digits:
        reference_digits = 0
    else:
        reference_digits = int(reference_digits)

    difference = actual_digits - reference_digits
    factor = 10 ** abs(difference)

    if difference < 0:
        return value * factor
    elif difference > 0:
        return value / factor
    else:
        return value


symbols = ["BTCUSD", "ETHUSD", "LTCUSD", "XRPUSD", "BCHUSD", "AAVEUSD", "ADAUSD", "ALGOUSD", "ATOMUSD", "AVAXUSD",
           "AXSUSD", "BNBUSD", "DASHUSD", "DOGEUSD", "DOTUSD", "FILUSD", "FTMUSD", "GRTUSD", "ICPUSD", "IOTAUSD",
           "LINKUSD", "LRCUSD", "MANAUSD", "MATICUSD", "NEARUSD", "SOLUSD", "UNIUSD", "ZECUSD", "ETCUSD", "TRXUSD",
           "AUS200", "UK100", "FRA40", "ESP35", "EUSTX50", "NAS100", "US30", "SPX500", "JPN225", "GER40", "XBRUSD",
           "XTIUSD", "XNGUSD", "XAGUSD", "XAUUSD", "XAUEUR"]

symbols_digits = {"BTCUSD": 2, "ETHUSD": 2, "LTCUSD": 2, "XRPUSD": 5,	"BCHUSD": 2, "AAVEUSD": 2, "ADAUSD": 4,
                  "ALGOUSD": 4, "ATOMUSD": 3,	"AVAXUSD": 3, "AXSUSD": 2, "BNBUSD": 2,	"DASHUSD": 2, "DOGEUSD": 5,
                  "DOTUSD": 3, "FILUSD": 3, "FTMUSD": 4, "GRTUSD": 4,	"ICPUSD": 3, "IOTAUSD": 4, "LINKUSD": 3,
                  "LRCUSD": 4, "MANAUSD": 4, "MATICUSD": 4, "NEARUSD": 3, "SOLUSD": 2, "UNIUSD": 3, "ZECUSD": 2,
                  "ETCUSD": 3, "TRXUSD": 5, "AUS200": 1, "UK100": 1, "FRA40": 1, "ESP35": 1, "EUSTX50": 1, "NAS100": 1,
                  "US30": 1, "SPX500": 1,	"JPN225": 0, "GER40": 1, "XBRUSD": 2, "XTIUSD": 2, "XNGUSD": 3,	"XAGUSD": 3,
                  "XAUUSD": 2, "XAUEUR": 2}

info = symbol_info("JPN225")
print(info)

tick_data = {}

for symbol in symbols:
    tick = symbol_info_tick(symbol)
    info = symbol_info(symbol)

    if not tick or not info:
        print(f"Could not retrieve tick or info for {symbol}.")
        continue

    if tick and info:
        # Determine the currency pair or other instrument
        if symbol[-3:] in ['USD', 'EUR', 'GBP', 'JPY', 'CAD', 'AUD', 'NZD', 'CHF']:
            quote_currency = symbol[-3:]
        else:
            quote_currency = info.currency_profit  # or info.currency_margin

        usd_rate = get_usd_rate(quote_currency)

        if not usd_rate:
            continue

        digits = info.digits
        spread = tick.ask - tick.bid  # This is the raw spread

        if digits != 0:
            spread_in_points = spread * (10 ** digits) # adjust raw spread to points
        else:
            spread_in_points = spread  # or any other default value

        spread_in_quote = spread  # Use raw spread for calculations
        spread_in_usd_per_contract = spread_in_quote * usd_rate * info.trade_contract_size
        spread_in_usd_per_contract = np.round(spread_in_usd_per_contract, 2)

        min_volume = info.volume_min
        max_volume = info.volume_max
        limit_volume = info.volume_limit
        contract_size = info.trade_contract_size
        margin = info.margin_initial

        swap_mode = info.swap_mode
        swap_long = info.swap_long if info.swap_mode != 0 else 0
        swap_short = info.swap_short if info.swap_mode != 0 else 0
        limit_volume = info.volume_limit
        trade_calc_mode = info.trade_calc_mode
        underlying_lot_amount_usd = tick.bid * contract_size * usd_rate
        underlying_lot_amount_usd = np.round(underlying_lot_amount_usd, 2)

        tick_data[symbol] = {
            "price": tick.ask,
            "digits": digits,
            "spread": spread,
            "spread_in_points": spread_in_points,  # added to the dictionary
            "contract_size": contract_size,
            "min_volume": min_volume,
            "max_volume": max_volume,
            "limit_volume": limit_volume,
            "margin": margin,
            "quote_currency": quote_currency,
            "usd_rate": usd_rate,
            "spread_in_usd_per_contract": spread_in_usd_per_contract,
            "swap_mode": swap_mode,
            "swap_long": swap_long,
            "swap_short": swap_short,
            "type": "",  # Empty by default
            "commission": "",  # Empty by default
            "underlying_lot_amount_usd": underlying_lot_amount_usd,
            "ib_commission_per_lot": "",  # Empty by default
            "trade_calc_mode": trade_calc_mode
        }

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Write headers to Excel
headers = ["Symbol", "Type", "Price", "Digits", "Spread", "Spread in Points", "Spread in Octa Points", "Spread in USD",
           "Contract Size", "Min Volume", "Max Volume", "Limit Volume", "Margin Buy", "Trade Calculation Mode",
           "Quote Currency", "USD rate", "Commission", "Swap mode", "Swap Long", "Swap Short",
           "Swap Long in Octa Points", "Swap Short in Octa Points", "Underlying lot amount USD",
           "IB Commission per lot"]

for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num).value = header

# Create a yellow fill for background color
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Write data to Excel
row_num = 2
for symbol in symbols:  # We loop through symbols list to make sure all symbols are captured
    data = tick_data.get(symbol, {})  # Use a default empty dictionary if symbol data is missing

    if symbol in ["AUS200", "XBRUSD", "XAGUSD"]:
        row_num += 1  # Increment row_num to leave an empty row

    ws.cell(row=row_num, column=1).value = symbol
    ws.cell(row=row_num, column=2).value = data.get("type", "")
    ws.cell(row=row_num, column=3).value = data.get("price", "")
    ws.cell(row=row_num, column=4).value = data.get("digits", "")
    ws.cell(row=row_num, column=5).value = data.get("spread", "")
    ws.cell(row=row_num, column=6).value = data.get("spread_in_points", "")
    # Adjusting spread_in_points
    actual_digits = data.get("digits", "")
    reference_digits = symbols_digits.get(symbol, actual_digits)  # default to actual_digits if symbol not found
    ws.cell(row=row_num, column=7).value = adjust_value_by_digits(data.get("spread_in_points", 0), actual_digits,
                                                                  reference_digits)

    ws.cell(row=row_num, column=8).value = data.get("spread_in_usd_per_contract", "")
    ws.cell(row=row_num, column=9).value = data.get("contract_size", "")
    ws.cell(row=row_num, column=10).value = data.get("min_volume", "")
    ws.cell(row=row_num, column=11).value = data.get("max_volume", "")
    ws.cell(row=row_num, column=12).value = data.get("limit_volume", "")
    ws.cell(row=row_num, column=13).value = data.get("margin", "")
    ws.cell(row=row_num, column=14).value = data.get("trade_calc_mode", "")
    ws.cell(row=row_num, column=15).value = data.get("quote_currency", "")
    ws.cell(row=row_num, column=16).value = data.get("usd_rate", "")
    ws.cell(row=row_num, column=17).value = data.get("commission", "")
    ws.cell(row=row_num, column=18).value = data.get("swap_mode", "")
    ws.cell(row=row_num, column=19).value = data.get("swap_long", "")
    ws.cell(row=row_num, column=20).value = data.get("swap_short", "")
    # Adjusting swap values
    ws.cell(row=row_num, column=21).value = adjust_value_by_digits(data.get("swap_long", 0), actual_digits,
                                                                   reference_digits)
    ws.cell(row=row_num, column=22).value = adjust_value_by_digits(data.get("swap_short", 0), actual_digits,
                                                                   reference_digits)

    ws.cell(row=row_num, column=23).value = data.get("underlying_lot_amount_usd", "")
    ws.cell(row=row_num, column=24).value = data.get("ib_commission_per_lot", "")

    # Apply the yellow fill to the appropriate cells
    ws.cell(row=row_num, column=2).fill = yellow_fill  # Type column
    ws.cell(row=row_num, column=17).fill = yellow_fill  # Commission column
    ws.cell(row=row_num, column=24).fill = yellow_fill  # IB Commission per lot column

    row_num += 1

for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:  # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)  # adding a little extra space
    ws.column_dimensions[column[0].column_letter].width = adjusted_width

# Save the workbook
wb.save("symbols_info.xlsx")

# Shutdown the connection to MetaTrader 5
shutdown()
